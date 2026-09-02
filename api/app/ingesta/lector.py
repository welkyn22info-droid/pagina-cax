"""Detecta TXT vs Excel y normaliza a un DataFrame con las columnas de
destino del esquema, dejando lo no reconocido en campos_extra (sección 9)."""
from __future__ import annotations

import datetime
import io
import re
import unicodedata

import pandas as pd
from openpyxl import load_workbook

from app.ingesta.esquemas import EsquemaInsumo

SEPARADORES_CANDIDATOS = ["|", ";", "\t", ","]


def _normalizar_encabezado(texto: str) -> str:
    """Sin mayúsculas, sin acentos, sin espacios sobrantes — para el mapeo
    flexible de encabezados que pide la sección 9."""
    texto = unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"\s+", " ", texto).strip().upper()


def _detectar_separador(contenido: bytes) -> str:
    primeras_lineas = contenido.decode("utf-8", errors="ignore").splitlines()[:5]
    muestra = "\n".join(primeras_lineas)
    conteos = {sep: muestra.count(sep) for sep in SEPARADORES_CANDIDATOS}
    mejor = max(conteos, key=conteos.get)
    if conteos[mejor] == 0:
        return r"\s+"  # ancho fijo / separado por espacios como último recurso
    return mejor


def _parsear_numero_colombiano(valor) -> float | None:
    """Punto como separador de miles y coma decimal cuando aparece, sin
    corromper valores que ya vienen en formato estándar (punto decimal)."""
    if valor is None:
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    texto = str(valor).strip()
    if texto == "" or texto.lower() in ("nan", "none", "null"):
        return None
    texto = texto.replace(" ", "")
    if "," in texto and "." in texto:
        # 1.234.567,89 -> formato colombiano completo
        if texto.rfind(",") > texto.rfind("."):
            texto = texto.replace(".", "").replace(",", ".")
        else:
            texto = texto.replace(",", "")
    elif "," in texto:
        # solo coma: es el decimal (1234,56)
        texto = texto.replace(".", "").replace(",", ".")
    # solo punto o ningún separador: se asume ya es formato estándar
    try:
        return float(texto)
    except ValueError:
        raise ValueError(f"'{valor}' no es un número válido")


def _parsear_fecha(valor):
    if valor is None or (isinstance(valor, float) and pd.isna(valor)):
        return None
    if isinstance(valor, datetime.datetime):
        return valor.date()
    if isinstance(valor, datetime.date):
        return valor
    # Intenta primero ISO (formato de la API y de flujos_pasivo en los
    # fixtures); si no calza, cae a dd/mm/aaaa (formato de la interfaz,
    # sección 12) sin que pandas emita la advertencia de ambigüedad.
    fecha = pd.to_datetime(valor, format="%Y-%m-%d", errors="coerce")
    if pd.isna(fecha):
        fecha = pd.to_datetime(valor, dayfirst=True, errors="coerce")
    if pd.isna(fecha):
        raise ValueError(f"'{valor}' no es una fecha válida")
    return fecha.date()


class ResultadoLectura:
    def __init__(self, df: pd.DataFrame, filas_leidas: int, encabezados_originales: list[str]):
        self.df = df
        self.filas_leidas = filas_leidas
        self.encabezados_originales = encabezados_originales


def _leer_excel_streaming(contenido: bytes, esquema: EsquemaInsumo) -> pd.DataFrame:
    """Lee un .xlsx con openpyxl en modo read_only: itera fila por fila en
    vez de materializar todo el libro en memoria de una vez, como hace
    pd.read_excel. Con archivos grandes (decenas de miles de filas) esto
    evita el pico de memoria y es notablemente más rápido — la regla de
    "Excel grandes" de la guía de arquitectura. Además conserva el tipo
    nativo de cada celda (número, fecha) en vez de forzar todo a texto."""
    libro = load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
    try:
        hoja = libro.worksheets[esquema.hoja] if isinstance(esquema.hoja, int) else libro[esquema.hoja]
        filas_iter = hoja.iter_rows(values_only=True)
        for _ in range(esquema.filas_encabezado):
            next(filas_iter, None)
        encabezados = [str(c).strip() if c is not None else "" for c in next(filas_iter, ())]
        filas = [fila for fila in filas_iter if any(v is not None for v in fila)]
    finally:
        libro.close()
    return pd.DataFrame(filas, columns=encabezados if encabezados else None)


def leer_archivo(nombre_archivo: str, contenido: bytes, esquema: EsquemaInsumo) -> ResultadoLectura:
    """Lee TXT o Excel según la extensión (verificando el contenido) y
    devuelve un DataFrame con las columnas de destino del esquema más
    campos_extra con lo que no se reconoce. No valida obligatoriedad —
    eso lo hace validador.py."""
    nombre_min = nombre_archivo.lower()

    if nombre_min.endswith(".xlsx"):
        crudo = _leer_excel_streaming(contenido, esquema)
    elif nombre_min.endswith(".xls"):
        # Formato binario legado: openpyxl no lo lee: cae al motor
        # tradicional de pandas (xlrd), sin streaming.
        crudo = pd.read_excel(
            io.BytesIO(contenido), sheet_name=esquema.hoja, header=esquema.filas_encabezado, dtype=str,
        )
    else:
        # Se autodetecta siempre a partir del contenido, aunque el esquema
        # declare un separador "habitual": los archivos varían entre
        # proveedores (sección 9), y fijarlo de antemano rompería
        # justo el caso que la autodetección existe para cubrir.
        separador = _detectar_separador(contenido)
        texto = contenido.decode("utf-8-sig", errors="ignore")
        crudo = pd.read_csv(
            io.StringIO(texto),
            sep=separador,
            engine="python",
            dtype=str,
            skipinitialspace=True,
            skiprows=esquema.filas_encabezado,
        )

    encabezados_originales = list(crudo.columns)
    mapa_normalizado_a_original = {_normalizar_encabezado(str(c)): c for c in crudo.columns}

    columnas_destino = {}
    columnas_usadas_originales: set[str] = set()
    for columna in esquema.columnas:
        col_original = None
        for candidato in columna.origen:
            candidato_norm = _normalizar_encabezado(candidato)
            if candidato_norm in mapa_normalizado_a_original:
                col_original = mapa_normalizado_a_original[candidato_norm]
                break
        if col_original is not None:
            columnas_destino[columna.destino] = crudo[col_original]
            columnas_usadas_originales.add(col_original)
        else:
            columnas_destino[columna.destino] = pd.Series([None] * len(crudo))

    salida = pd.DataFrame(columnas_destino)

    columnas_sobrantes = [c for c in crudo.columns if c not in columnas_usadas_originales]
    if columnas_sobrantes:
        salida["campos_extra"] = crudo[columnas_sobrantes].apply(
            lambda fila: {k: v for k, v in fila.items() if pd.notna(v)}, axis=1
        )
    else:
        salida["campos_extra"] = [{} for _ in range(len(crudo))]

    return ResultadoLectura(df=salida, filas_leidas=len(crudo), encabezados_originales=encabezados_originales)
